from flask import Flask,render_template, request, url_for, redirect, send_from_directory
from flask_sqlalchemy import SQLAlchemy
import os
import os.path
from openpyxl import load_workbook, Workbook
import time
import socket

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data/users.db'
db = SQLAlchemy(app)

class Users(db.Model): # создание базы данных в программе
    id = db.Column(db.Integer, primary_key=True)
    Username = db.Column(db.String(64), nullable=False)
    Password = db.Column(db.String(16), nullable=False)
    Root = db.Column(db.String(6), nullable=False)

    def __repr__(self):
        return '<{} {}>'.format(self.username) 

@app.route('/admin/<username>', methods=['post', 'get'])
def admin(username):

    return render_template('admin.html', username = username, root = 'admin')

@app.route ('/user/<username>', methods=['post', 'get'])
def user(username):

    return render_template('user.html', username = username, root = 'user')

@app.route ('/come/<username>/<root>') # Отметка пользователя об уходе
def come(username,root):
    wb = load_workbook('data/data.xlsx')
    ws = wb.active
    rows = ws.max_row
    log = open(os.path.join(os.path.dirname(__file__), 'data', 'log.txt'), 'a')
    for i in range(1, rows + 1):
        if ws.cell(i,1).value == username and ws.cell(i,3).value == None: 
            if root == 'admin': # Проверка роли пользователя, и на ошибку пользователя на повторное нажатие
                log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' ошибка в отметке прихода, пользователь:' + username + '\n')
                log.close
                return render_template('admin.html', username = username, message  = 'come')
            else:
                log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' ошибка в отметке прихода, пользователь:' + username + '\n')
                log.close
                return render_template('user.html', username = username, message  = 'come')
    ws.cell(rows + 1,1).value = username # запись в excel
    ws.cell(rows + 1,2).value = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    wb.save('data/data.xlsx')
    if root == 'admin': # Проверка роли пользователя, и запись времени прихода, время системное
        log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' отметка прихода, пользователь:' + username + '\n')
        log.close
        return render_template('admin.html', username = username, message = 'ok') 
    else:
        log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' отметка прихода, пользователь:' + username + '\n')
        log.close
        return render_template('user.html', username = username, message = 'ok') 

@app.route ('/out/<username>/<root>')
def out(username,root): # Отметка пользователя об уходе
    wb = load_workbook('data/data.xlsx')
    ws = wb.active
    rows = ws.max_row
    log = open(os.path.join(os.path.dirname(__file__), 'data', 'log.txt'), 'a')
    for i in range(1, rows + 1):
        if ws.cell(i, 1).value == username and ws.cell(i, 3).value == None: 
            ws.cell(i,3).value = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) # запись в excel
            wb.save('data/data.xlsx')
            if root == 'admin': # Проверка роли пользователя, и запись времени ухода, время системное
                log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' отметка ухода, пользователь:' + username + '\n')
                log.close
                return render_template('admin.html', username = username, message = 'ok')
            else:
                log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' отметка ухода, пользователь:' + username + '\n')
                log.close
                return render_template('user.html', username = username, message = 'ok')
    if root == 'admin': # Проверка роли пользователя, и на ошибку пользователя на случайное нажатие
        log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' ошибка в отметке ухода, пользователь:' + username + '\n')
        log.close
        return render_template('admin.html', username = username, message = 'out')
    else:
        log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' ошибка в отметке ухода, пользователь:' + username + '\n')
        log.close
        return render_template('user.html', username = username, message = 'out')

@app.route ('/register/<username>', methods=['post', 'get'])
def register(username): # Создание нового пользователя
    if request.method == 'POST':
        useradmin = username
        username = request.form.get('username')
        password = request.form.get('password')
        root = request.form.get('root')
        username = username.lower()
        root = root.lower()
        log = open(os.path.join(os.path.dirname(__file__), 'data', 'log.txt'), 'a')

        user = Users(Username=username, Password=password, Root=root)
        try: 
            db.session.add(user) # Записываем в базу данных
            db.session.commit()
            return render_template('admin.html', username = useradmin, message = 'reg')
        except:
            log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' регистрация нового пользователя не успешна\n')
            log.close
            return 'Error'
    return render_template('register.html')

@app.route ('/filter', methods=['post', 'get'])
def filter(): # настраиваем фильтр по фамилии
    filter = ''
    if request.method == 'POST': 
        log = open(os.path.join(os.path.dirname(__file__), 'data', 'log.txt'), 'a')
        filter = request.form.get('filter')
        filter = filter.lower()
        wb = load_workbook('data/data.xlsx')
        ws = wb.active
        rows = ws.max_row
    
        fb = Workbook()
        fb.create_sheet("Filtered")
        fs = fb.active
        j = 1
        c = 0

        for i in range(1, rows + 1):
            if ws.cell(i,1).value == filter: # записываем фильтр в отдельный файл
                fs.cell(j,1).value = ws.cell(i,1).value
                fs.cell(j,2).value = ws.cell(i,2).value
                fs.cell(j,3).value = ws.cell(i,3).value
                j+=1
                c = 1
        fb.save('data/test_fil.xlsx')
        if c == 1: # Скачиваем файл в случае найденных пользователей
            log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' попытка скачать отчет с фильтром '+ filter + '\n')
            log.close
            c = 0
            return send_from_directory('data','test_fil.xlsx')
        else:
            log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' пользователь не найден '+ filter + '\n')
            log.close
            return render_template('filter.html', message = 'no')
    return render_template('filter.html')
    
@app.route ('/download/<type_d>', methods=['post', 'get'])
def download(type_d):
    log = open(os.path.join(os.path.dirname(__file__), 'data', 'log.txt'), 'a')
    if type_d == 'logs': #  скачиваем логи
        try:
            filename = 'log.txt'
            log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' попытка скачать лог\n')
            log.close            
            return send_from_directory('data',filename,as_attachment = True)
        except:
            return('error')
    if type_d == 'data': # скачиваем данные пользователей о нахождении на работе
        try:
            log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' попытка скачать отчет\n')
            log.close
            return send_from_directory('data','data.xlsx')
        except:
            return('error')
    return render_template('download.html', type_d = 'none')

@app.route('/', methods=['post', 'get'])
def index():
    log = open(os.path.join(os.path.dirname(__file__), 'data', 'log.txt'), 'a')
    log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' кто-то зашел на сервер\n') # запись лога
    message = ''
    username = ''
    password = ''
    if os.path.isfile('data/data.xlsx') == False: #Создаем файл эксель, если он отсутствует
        wb = Workbook()
        wb.create_sheet("Mysheet")
        ws = wb.active
        ws.cell(1,1).value = 'Имя пользователя'
        ws.cell(1,2).value = 'Время прихода'
        ws.cell(1,3).value = 'Время ухода'
        wb.save('data/data.xlsx')
    if request.method == 'POST':
        username = request.form.get('username')  # запрос к данным формы входа
        password = request.form.get('password')
        username = username.lower()
        user = Users.query.all()
        for u in user:
            if username == u.Username and password == u.Password: # проверка пользователя
                if u.Root == 'admin': # проверка прав пользователя
                    log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' Пользователь: ' + username + ' права: администратор авторизовался\n')
                    log.close
                    return redirect(url_for('admin', username = u.Username, root = 'admin'))
                else:
                    log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' Пользователь: ' + username + ' права: пользователь авторизовался\n')
                    log.close
                    return redirect(url_for('user', username = u.Username, root = 'user'))
            else:
                log.write(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' Пользователь: ' + username + ' неверное имя или пароль\n')
                log.close
                message = "0"
    return render_template('index.html', message=message)


if __name__ == "__main__":
    app.run(debug='true', host='127.0.0.1', port=8000)